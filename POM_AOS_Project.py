from random import randint
from unittest import TestCase

from selenium.webdriver.chrome.service import Service
from selenium import webdriver

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from Classes.TopBar_Class import Topbar
from Classes.Homepage_Class import Homepage
from Classes.Category_Class import Category
from Classes.Product_Class import Product
from Classes.Cart_Class import Cart
from Classes.OrderPayment_Class import OrderPayment
from Classes.CreateAccount_Class import CreateAccount
from Classes.MyOrders_Class import MyOrders


class test_AOS_Website(TestCase):

    def setUp(self):
        # loading the excel file 'data.xlsx'.
        self.wb = load_workbook('data.xlsx')
        self.ws = self.wb.active
        self.headers_row = 1
        self.results_row = 21
        self.field_col = self.excel_get_fields_col()

        # setting a variable for 'tearDown' method
        self.did_test_pass = False

        self.service = Service(r'D:\QA_Course\webdrivers\chromedriver.exe')
        self.driver = webdriver.Chrome(service=self.service)
        self.driver.get("https://www.advantageonlineshopping.com/#/")
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()

        self.topbar = Topbar(self.driver)
        self.homepage = Homepage(self.driver)
        self.category = Category(self.driver)
        self.product = Product(self.driver)
        self.cart = Cart(self.driver)
        self.order_payment = OrderPayment(self.driver)
        self.create_account = CreateAccount(self.driver)
        self.my_orders = MyOrders(self.driver)
        if self.topbar.is_logged_in():
            self.topbar.click_sign_out()

    def tearDown(self):
        # closing the driver:
        self.topbar.click_aos_logo()
        self.driver.close()

        test_num = self.str_to_num(self._testMethodName)
        self.excel_write_results(test_num, self.did_test_pass)

    ## IN-CLASS METHODS

    def add_specific_product(self, category: str, product_id: int, quantity: int):
        self.topbar.click_aos_logo()
        self.homepage.click_specific_category(category)
        self.category.click_specific_product(product_id)
        self.product.set_quantity(quantity)
        product_details = {
                'name': self.product.get_name(),
                'quantity': self.str_to_num(self.product.get_quantity()),
                'color': self.product.get_color(),
                'price': self.str_to_num(self.product.get_price())
            }
        self.product.add_to_cart()
        return product_details

    # receives a string and returns a number (int or float)
    def str_to_num(self, stri: str):
        """
        :param stri: string including numeric characters and no more than one dot (.)
        :return: that string as a number type (int/float), depends on the existence of a dot (.)
        """
        # if received a number, return it as it is:
        if type(stri) == int or type(stri) == float:
            return stri

        is_float = False
        for char in stri:
            if not char.isnumeric():
                if char == '.':
                    is_float = True
                    continue
                stri = stri.replace(char, '')
        if is_float:
            return float(stri)
        try:
            return int(stri)
        except:
            return ''

    # returns the column char (capital char) in the Excel file that represents 'field'
    def excel_get_fields_col(self):
        """
        :functionality: searches in the Excel file for the column that represents 'field'
        :return: the column as capital char
        """
        for col in range(2, 13):
            char = get_column_letter(col)
            header_value = self.ws[char + str(self.headers_row)].value
            if header_value == 'field':
                return char

    # receives a test number and return its matching column in the Excel file
    def excel_get_test_col(self, test_num):
        """
        :param test_num: test number
        :return: the corresponding column in the Excel file
        """
        # looping through the 'Headers' row to find the one matching the test_num
        test_col = ''
        for col in range(2, 13):
            char = get_column_letter(col)
            header_value = self.ws[char + str(self.headers_row)].value
            if self.str_to_num(header_value) == test_num:
                test_col = char
                break;
        return test_col

    # Excel: receives a field value (str) and the test number (int) and return a list of all the cells' values matching the given field and test number.
    def excel_get_values_lst(self, field, test_num):
        """
        :param field: field value (str), as specified in the Excel file "data.xlsx", workbook 'Data', column B
        :param test_num: the test number (int)
        :functionality: gets the Excel's cells' values matching the field and the test number.
        :return: list of all the matching cells' values.
                if there is only one value- returns it as a string.
        """
        test_col = self.excel_get_test_col(test_num)

        # looping through all of the rows to find the one matching the 'field'
        field_lst = []
        for row in range(1, 21):
            if self.ws[self.field_col + str(row)].value == field:
                # adding the cells matching both the 'test_num' and the 'field' to a list (field_lst)
                field_value = self.ws[test_col + str(row)].value
                if field_value is not None:
                    field_lst.append(field_value)

        if len(field_lst) == 1:
            return field_lst[0]
        return field_lst

    # for tearDown- writes for test_num if id did_pass or not, in the corresponding Excel file's cell
    def excel_write_results(self, test_num, did_pass):
        """
        :param test_num: test number
        :param did_pass: boolean value for whether the test passed or not
        :functionality: writes the result to the corresponding Excel file's cell
        :return: None
        """
        char = self.excel_get_test_col(test_num)
        cell = self.ws[char + str(self.results_row)]
        if did_pass:
            cell.value = 'V'
        else:
            cell.value = 'X'
        self.wb.save('data.xlsx')

    ## TESTS FUNCTIONS

    def test_1(self):
        """
        לאחר בחירה של לפחות שני מוצרים, בכמויות שונות (לדוגמה 2 יחידות של מוצר ראשון ו-3 יחידות של מוצר שני),
        יש לבדוק שכמות המוצרים הסופית מופיעה נכון ומדוייק בחלונית עגלת הקניות מצד ימין למעלה של המסך.
        """
        # pulling data from 'data.xlsx'
        test_num = 1
        categories = self.excel_get_values_lst('category', test_num)
        quantities = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)
        sum_quantities = 0
        for i in range(len(quantities)):
            self.add_specific_product(categories[i], ids[i], quantities[i])
            # calculating the total quantities from the excel file:
            sum_quantities += quantities[i]
        # getting the total quantities from the cart icon:
        cart_products_details = self.topbar.get_total_items()

        self.assertEqual(cart_products_details, sum_quantities)

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_2(self):
        """
        לאחר בחירה של 3 מוצרים, בכמויות שונות,
        יש לבדוק שהמוצרים מופיעים נכון:
        כולל שם, צבע, כמות ומחיר בחלונית עגלת הקניות בצד ימין למעלה של המסך
        """
        # pulling data from 'data.xlsx'
        test_num = 2
        categories = self.excel_get_values_lst('category', test_num)
        quantities = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)
        products_details = []
        for i in range(len(quantities)):
            products_details.append(self.add_specific_product(categories[i], ids[i], quantities[i]))

        # getting items details from the pop-up cart
        cart_products_details = self.topbar.get_items_from_popup_cart()

        # first test- equality of the two lists' length
        self.assertEqual(len(products_details), len(cart_products_details))

        # second test- checking the correctness of the product's details in the popup cart.
        for i in range(len(products_details)):
            self.assertEqual(products_details[i]['name'], cart_products_details[i]['name'])
            self.assertEqual(self.str_to_num(products_details[i]['quantity']), self.str_to_num(cart_products_details[i]['quantity']))
            self.assertEqual(products_details[i]['color'], cart_products_details[i]['color'])
            # calculating each product's price with its quantity
            product_price = products_details[i]['quantity'] * products_details[i]['price']
            self.assertEqual(product_price, self.str_to_num(cart_products_details[i]['price']))

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_3(self):
        """
        לאחר בחירה של לפחות שני מוצרים והסרה של מוצר אחד ע"י שימוש בחלונית עגלת הקניות למעלה מימין,
        יש לבדוק שהמוצר אכן נעלם מחלונית העגלה
        """
        # pulling data from 'data.xlsx'
        test_num = 3
        categories = self.excel_get_values_lst('category', test_num)
        quantities = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)
        for i in range(len(quantities)):
            self.add_specific_product(categories[i], ids[i], quantities[i])

        popup_cart_before = self.topbar.get_items_names_from_popup_cart()

        # choosing random product to remove
        del_index = randint(0, len(popup_cart_before) - 1)

        self.topbar.remove_item_from_popup_cart(del_index)

        popup_cart_after = self.topbar.get_items_names_from_popup_cart()

        # check if the removed item is still in the cart:
        if popup_cart_before[del_index] in popup_cart_after:
            self.fail("The removed item is still in the cart")

        # compare amount of cart items after removing
        self.assertEqual(len(popup_cart_after), len(popup_cart_before) - 1)

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_4(self):
        """
        לאחר בחירה של מוצר כלשהו ומעבר למסך עגלת הקניות ע"י לחיצה על לחצן עגלת הקניות בצד ימין למעלה,
        יש לוודא מעבר לעמוד עגלת הקניות
        (ע"י בדיקת הופעת הטקסט:
        "Shopping cart"
        למעלה משמאל)
        """
        # getting data from 'data.xlsx'
        test_num = 4
        category = self.excel_get_values_lst('category', test_num)
        prod_id = self.excel_get_values_lst('id', test_num)

        # going to the product's page:
        self.homepage.click_specific_category(category)
        self.category.click_specific_product(prod_id)

        self.topbar.click_cart()
        current_page = self.topbar.get_page()
        self.assertEqual('SHOPPING CART', current_page)

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_5(self):
        """
        לאחר בחירה של 3 מוצרים בכמויות שונות ומעבר לעמוד עגלת הקניות, יש לבדוק שהסכום הכולל של ההזמנה תואם
        את מחירי המוצרים והכמויות שהוזמנו,
        עפ"י סיכום המחירים שהופיעו *בעת בחירת המוצרים*.
        בטסט זה יש להדפיס בצורה ברור, עבור כל מוצר בעגלת הקניות:
        שם המוצר, כמות המוצר, מחיר המוצר
        """
        # pulling data from 'data.xlsx'
        test_num = 5
        categories = self.excel_get_values_lst('category', test_num)
        quantities = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)
        added_products = []
        for i in range(len(quantities)):
            added_products.append(self.add_specific_product(categories[i], ids[i], quantities[i]))

        total_price = 0
        for product in added_products:
            total_price += product['price'] * product['quantity']
            print(f"===== Product {added_products.index(product)+1}")
            print(f"Name:  {product['name']}")
            print(f"Quantity:  {product['quantity']}")
            print(f"Price:  {product['price']}")
        self.topbar.click_cart()
        cart_total = self.str_to_num(self.cart.get_total())
        print("=====\ncart_total: ", cart_total, "\ntotal_price: ", total_price)
        self.assertEqual(total_price, cart_total)

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_6(self):
        """
        לאחר בחירה של שני מוצרים לפחות, לעבור לעמוד עגלת הקניות ולבצע שני שינויים בכמויות של שני המוצרים.
        יש לבדוק שהשינויים משתקפים בעמוד עגלת הקניות.
        """
        # pulling data from 'data.xlsx':
        test_num = 6
        categories = self.excel_get_values_lst('category', test_num)
        quants_str = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)
        added_products = []
        new_quantities = []
        for i in range(len(categories)):
            # separating between the initial quantities to the new quantities information
            quants_sep = quants_str[i].split('->')
            new_quantities.append(self.str_to_num(quants_sep[1]))
            added_products.append(self.add_specific_product(categories[i], ids[i], self.str_to_num(quants_sep[0])))

        items_amount = len(self.topbar.get_items_elems_from_popup_cart())
        self.topbar.click_cart()

        # changing the quantities to all products
        for i in range(items_amount - 1):
            self.cart.click_edit(i)
            self.product.set_quantity(new_quantities[i])
            self.product.add_to_cart()

        # getting the new information from the cart
        cart_new_quantities = []
        for quantity in self.cart.get_products_quantities():
            cart_new_quantities.append(self.str_to_num(quantity))

        # comparing the new quantities information.
        for i in range(len(cart_new_quantities)):
            self.assertEqual(new_quantities[i], cart_new_quantities[i])

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_7(self):
        """
        לאחר בחירה של מוצר מסוג טאבלט,
        יש לחזור אחורה ולבדוק שחזרנו למסך מוצרי ה- טאבלטים,
        ושוב לחזור אחורה ולבדוק שחזרנו למסך הראשי
        """
        # pulling data from 'data.xlsx'
        test_num = 7
        category = self.excel_get_values_lst('category', test_num)
        quantity = self.excel_get_values_lst('quantity', test_num)
        prod_id = self.excel_get_values_lst('id', test_num)

        # adding the product to the cart
        self.add_specific_product(category, prod_id, quantity)

        # going back and checking the page path
        self.topbar.go_back()
        self.assertEqual(self.topbar.get_page(), category.upper())

        # going back and checking the page path
        self.topbar.go_back()
        self.assertEqual(self.topbar.get_page(), 'Home')

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_8(self):
        """
        לאחר בחירת מוצרים כלשהם לקנייה, לבצע checkout,
        למלא פרטי משתמש חדש ,
        לבצע תשלום באמצעות SafePay,
        לבדוק שהתשלום בוצע בהצלחה,
        לבדוק שעגלת הקניות ריקה,
        ושההזמנה נמצאת ברשימת ההזמנות של המשתמש.
        """
        ## check if the user exists (try to log in) and if yes- delete it/ change the username.
        # pulling data from 'data.xlsx'

        test_num = 8
        categories = self.excel_get_values_lst('category', test_num)
        quantities = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)

        # adding the products to the cart
        for i in range(len(quantities)):
            self.add_specific_product(categories[i], ids[i], quantities[i])

        # checkout
        self.topbar.click_cart()
        self.cart.click_checkout()
        self.order_payment.click_registration()

        # creating a new account
        new_username = self.excel_get_values_lst('new username', test_num)
        new_password = self.excel_get_values_lst('new password', test_num)
        new_email = self.excel_get_values_lst('new email', test_num)
        self.create_account.set_username(new_username)
        self.create_account.set_password(new_password, True)
        self.create_account.set_email(new_email)
        self.create_account.check_i_agree()
        self.create_account.click_register()

        # dealing with checkbox not marked
        try:
            self.order_payment.click_next()
        except:
            print("user already exists")

        # entering 'Safepay' account details
        safepay_username = self.excel_get_values_lst('safepay username', test_num)
        safepay_password = self.excel_get_values_lst('safepay password', test_num)
        self.order_payment.set_safepay_username(safepay_username)
        self.order_payment.set_safepay_password(safepay_password)
        self.order_payment.click_pay_now_safepay()

        # getting the order number after it's been successfully committed
        order_number = self.order_payment.get_order_number()

        # checking if the cart is empty
        self.topbar.click_cart()
        if not self.cart.is_cart_empty():
            self.fail("The cart is not empty after committing an order.")

        # getting all of the orders numbers from 'My Orders' page.
        self.topbar.click_my_orders()
        orders_numbers = self.my_orders.get_orders_numbers()

        # checking if the recent order exists in 'My Orders' page.
        if order_number not in orders_numbers:
            self.fail(f"The order {order_number} is not in 'My Orders' page.")

        # deleting the user for re-running the test:
        print("@@@ DELETING THE ACCOUNT. THIS IS NOT PART OF THE TEST @@@")
        self.topbar.delete_logged_user()

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_9(self):
        """
        לאחר בחירת מוצרים כלשהם לקניה, לבצע checkout,
        להחתבר למשתמש קיים,
        לבצע תשלום באמצעות כרטיס אשראי,
        לבדוק שעגלת הקניות ריקה
        ושההזמנה נמצאת ברשימה ההזמנות של המשתמש
        """
        ## check if the user exists (try to log in) and if not- create it.
        test_num = 9

        # pulling data from 'data.xlsx':
        categories = self.excel_get_values_lst('category', test_num)
        quantities = self.excel_get_values_lst('quantity', test_num)
        ids = self.excel_get_values_lst('id', test_num)
        products_details = []
        for i in range(len(quantities)):
            products_details.append(self.add_specific_product(categories[i], ids[i], quantities[i]))

        self.topbar.click_cart()
        self.cart.click_checkout()

        # logging in to an existing account:
        aos_username = self.excel_get_values_lst('existing username', test_num)
        aos_password = self.excel_get_values_lst('existing password', test_num)
        self.order_payment.set_username(aos_username)
        self.order_payment.set_password(aos_password)
        self.order_payment.click_login()

        self.order_payment.click_next()

        # selecting 'MasterCard' payment method
        self.order_payment.select_mastercard()

        # getting data from the Excel file
        card_number = self.excel_get_values_lst('mastercard card number', test_num)
        cvv_number = self.excel_get_values_lst('mastercard cvv number', test_num)
        cardholder_name = self.excel_get_values_lst('mastercard cardholder name', test_num)

        # setting the details to the form
        self.order_payment.set_mastercard_cardholder_name(cardholder_name)
        self.order_payment.set_mastercard_cvv_number(cvv_number)
        self.order_payment.set_mastercard_card_number(card_number)

        self.order_payment.click_pay_now_mastercard()

        # getting the order number after it's been successfully committed
        order_number = self.order_payment.get_order_number()

        # checking if the cart is empty
        self.topbar.click_cart()
        if not self.cart.is_cart_empty():
            self.fail("The cart is not empty after committing an order.")

        # getting all of the orders numbers from 'My Orders' page.
        self.topbar.click_my_orders()
        orders_numbers = self.my_orders.get_orders_numbers()

        # checking if the recent order exists in 'My Orders' page.
        if order_number not in orders_numbers:
            self.fail(f"The order {order_number} is not in 'My Orders' page.")

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

    def test_10(self):
        """
        לבדוק תהליכי התחברות והתנתקות:
        לבצע התחברות למערכת באמצעות משתמש קיים ולוודא שהחיבור הצליח
        לבצע התנתקות ולוודא שההתנתקות הצליחה
        """
        ## check if the user exists (try to log in) and if not- create it.
        test_num = 10
        if self.topbar.is_logged_in():
            self.topbar.click_sign_out()
        aos_username = self.excel_get_values_lst('existing username', test_num)
        aos_password = self.excel_get_values_lst('existing password', test_num)
        self.topbar.login(aos_username, aos_password)
        self.assertTrue(self.topbar.is_logged_in())
        self.topbar.click_sign_out()
        self.assertFalse(self.topbar.is_logged_in())

        # telling the tearDown() method that this test passed
        self.did_test_pass = True

